import pytest
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.alert import Alert
import time
'''
                *****************************
                | TEST DRIVEN PYTEST MODULE |
                *****************************

 --> Demonstrates Chrome Driver Manager
 --> Demonstrates using Excel file for test driver
 --> Demonstrates iterating through Excel rows and test cases
 --> Demonstrates handling null entry
 --> Demonstrates Pytest, Pytest HTML reporting, Pytest Logging, Pytest report image inserts
 --> Demonstrates Assertions, Pytest Parametrizing, and error handling
 --> Demonstrates Alert handling
 --> Demonstrate Web Page images output support

'''
@pytest.fixture(scope="module")
def setup():
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
    driver.maximize_window()
    yield driver
    # driver.quit()

def load_test_data():
    test_data = []
    workbook = load_workbook(filename="test_data.xlsx")
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        test_data.append(row)

    return test_data

def check_for_alert(driver):
    try:
        alert = driver.switch_to_alert()
        alert.dismiss()
    except:
        pass

def check_password(password):
    if password == "None":
        password = None
    return str(password)

@pytest.mark.parametrize("username, password, expected_result", load_test_data())
def test_login(setup, username, password, expected_result):
    driver = setup
    driver.switch_to.new_window('tab')

    driver.get("https://www.saucedemo.com/")
    chg_password = check_password(password)
    driver.find_element(By.ID, "user-name").send_keys(username)
    if chg_password == "None":
        driver.find_element(By.ID, "password").send_keys('')
    else:
        driver.find_element(By.ID, "password").send_keys(chg_password)
    driver.find_element(By.ID, "login-button").click()


    time.sleep(2)
    if expected_result == "success":
        check_for_alert(driver)
        element = driver.find_element(By.CSS_SELECTOR, ".title")
        time.sleep(1)
        assert element.is_displayed() == True
        assert element.text == "Products"
        print("Sauce Demo Page Displayed: ", element.text)
    else:
        element = driver.find_element(By.TAG_NAME, "h3")
        time.sleep(1)
        assert element.is_displayed() == True
        assert "Epic" in element.text
        print("Sauce Demo Error Condition: ", element.text)


if __name__ == "__main__":
    pytest.main(["-v", "--html=report.html", "--log-cli-level=INFO"])
